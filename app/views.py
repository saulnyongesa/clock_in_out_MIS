import base64
import itertools
from io import BytesIO
import datetime
import json
from datetime import datetime as t2
from datetime import timedelta
from django.utils.timezone import make_aware
import cv2
import face_recognition
import numpy as np
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.db.models import Q
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from reportlab.lib.pagesizes import landscape, A4, portrait
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

from base.forms import *


# Create your views here.
def dashboard(request):
    return render(request, 'clock/clock_home_face.html')


# =======================Unit Views===========================
@login_required(login_url='clock-face-url')
def units_list(request):
    if request.user.is_superuser:
        units = Unit.objects.all()
        context = {
            'units': units,
        }
        return render(request, 'clock/units.html', context)
    else:
        return render(request, 'Error.html')


def unit_signup(request):
    if request.user.is_superuser:
        form = UnitSignupForm()
        if request.method == "POST":
            form = UnitSignupForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Unit added successfully')
                return redirect('unit-signup-url')
        context = {
            'form': form,
        }
        return render(request, 'clock/unit_signup.html', context)
    else:
        return render(request, 'Error.html')


def unit_edit(request, pk):
    if request.user.is_superuser:
        unit = Unit.objects.get(id=pk)
        form = UnitSignupForm(instance=unit)
        if request.method == "POST":
            form = UnitSignupForm(request.POST, instance=unit)
            if form.is_valid():
                form.save()
                messages.success(request, 'Unit updated successfully')
                return redirect('unit-signup-url')
        context = {
            'form': form,
            'unit': unit,
        }
        return render(request, 'clock/unit_edit.html', context)
    else:
        return render(request, 'Error.html')


# =======================Unit Views End===========================


# ======================Trainer data ===================================
def trainer_signup(request):
    if request.user.is_superuser:
        form = TrainerSignupForm()
        if request.method == "POST":
            form = TrainerSignupForm(request.POST, request.FILES)
            if form.is_valid():
                trainer = form.save(commit=False)
                uploaded_photo = request.FILES.get('photo')
                if uploaded_photo:
                    file_name = default_storage.save(uploaded_photo.name, uploaded_photo)
                    file_path = default_storage.path(file_name)
                    with open(file_path, 'rb') as f:
                        image_data = f.read()
                    nparr = np.frombuffer(image_data, np.uint8)
                    frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                    face_encodings = face_recognition.face_encodings(frame)
                    if face_encodings:
                        face_encoding = face_encodings[0]
                        trainer.face_encoding = ','.join(map(str, face_encoding))
                        trainer.save()
                        messages.success(request, 'Trainer added successfully')
                        return redirect('trainer-signup-url')
                    else:
                        messages.error(request, 'No face detected in the uploaded photo.')
                else:
                    messages.error(request, 'Please upload a photo.')

        context = {
            'form': form,
        }
        return render(request, 'clock/trainer_signup.html', context)
    else:
        return render(request, 'Error.html')


def trainer_edit(request, pk):
    if request.user.is_superuser:
        trainer = Trainer.objects.get(id=pk)
        form = TrainerSignupForm(instance=trainer)
        if request.method == "POST":
            form = TrainerSignupForm(request.POST, request.FILES, instance=trainer)
            if form.is_valid():
                trainer = form.save(commit=False)
                uploaded_photo = request.FILES.get('photo')
                if uploaded_photo:
                    file_name = default_storage.save(uploaded_photo.name, uploaded_photo)
                    file_path = default_storage.path(file_name)
                    with open(file_path, 'rb') as f:
                        image_data = f.read()
                    nparr = np.frombuffer(image_data, np.uint8)
                    frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                    face_encodings = face_recognition.face_encodings(frame)
                    if face_encodings:
                        face_encoding = face_encodings[0]
                        trainer.face_encoding = ','.join(map(str, face_encoding))
                        trainer.save()
                        messages.success(request, 'Trainer details edited successfully')
                        return redirect('trainer-signup-url')
                    else:
                        messages.error(request, 'No face detected in the uploaded photo.')
                else:
                    messages.error(request, 'Please upload a photo.')

        context = {
            'form': form,
            'trainer': trainer
        }
        return render(request, 'clock/trainer_edit.html', context)
    else:
        return render(request, 'Error.html')


def assign_trainer_to_unit(request):
    if request.user.is_superuser:
        form = TrainerAssignToUnitForm()
        if request.method == "POST":
            form = TrainerAssignToUnitForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Trainer Assigned to Unit successfully')
                return redirect('assign-trainer-to-unit-url')
        context = {
            'form': form,
        }
        return render(request, 'clock/trainer_assign_unit.html', context)
    else:
        return render(request, 'Error.html')


def assign_trainer_to_unit_from_trainer_list(request, pk):
    if request.user.is_superuser:
        trainer = Trainer.objects.get(id=pk)
        units = Unit.objects.all()
        if request.method == 'POST':
            unit_id = request.POST.get('unit_id')
            unit = Unit.objects.get(id=int(unit_id))
            try:
                if TrainerUnit.objects.get(
                        trainer_id=trainer,
                        unit=unit
                ):
                    messages.error(request, 'Trainer Already Assigned this unit')
                    return redirect('trainer-list-url')
            except TrainerUnit.DoesNotExist:
                TrainerUnit.objects.create(
                    trainer=trainer,
                    unit=unit
                )
                messages.success(request, trainer.name + " assigned " + unit.unit_name + ' (' + unit.unit_code + ')')
                return redirect('trainer-list-url')
        return render(request, 'clock/trainer_assign_unit_from_trainer_list.html', {'units': units, 'trainer': trainer})
    else:
        return render(request, 'Error.html')


def remove_assigned_unit(request, pk):
    if request.user.is_superuser:
        assigned_unit = TrainerUnit.objects.get(id=pk)
        assigned_unit.delete()
        messages.success(request, 'Unit removed')
        return redirect('trainer-list-url')
    else:
        return render(request, 'Error.html')


def trainer_status(request, pk):
    if request.user.is_superuser:
        trainer = Trainer.objects.get(id=pk)
        trainer.is_active = not trainer.is_active
        trainer.save()
        messages.success(request, 'Trainer active status changed')
        return redirect('trainer-list-url')
    else:
        return render(request, 'Error.html')


def trainer_list(request):
    if request.user.is_superuser:
        list_choice = 0
        units = TrainerUnit.objects.all()
        trainers = Trainer.objects.all()
        if request.method == 'POST':
            choice = request.POST.get('choice')
            if int(choice) == 0:
                return redirect('trainer-list-url')
            elif int(choice) == 1:
                return redirect('trainer-list-active-url')
            elif int(choice) == 2:
                return redirect('trainer-list-inactive-url')
        return render(request, 'clock/trainer_list.html',
                      {
                          'units': units,
                          'trainers': trainers,
                          'list_choice': list_choice})
    else:
        return render(request, 'Error.html')


def trainer_list_active_only(request):
    if request.user.is_superuser:
        list_choice = 1
        units = TrainerUnit.objects.all()
        trainers = Trainer.objects.filter(
            is_active=True
        )
        return render(request, 'clock/trainer_list.html',
                      {
                          'units': units,
                          'trainers': trainers,
                          'list_choice': list_choice})
    else:
        return render(request, 'Error.html')


def trainer_list_inactive_only(request):
    if request.user.is_superuser:
        list_choice = 2
        units = TrainerUnit.objects.all()
        trainers = Trainer.objects.filter(
            is_active=False
        )
        return render(request, 'clock/trainer_list.html',
                      {
                          'units': units,
                          'trainers': trainers,
                          'list_choice': list_choice})
    else:
        return render(request, 'Error.html')


# ======================Trainer data End===================================


# Student information=============================================================================
def student_signup(request):
    if request.user.is_superuser:
        form = StudentSignupForm()
        if request.method == "POST":
            form = StudentSignupForm(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                messages.success(request, 'Student added successfully')
                return redirect('student-signup-url')
        context = {
            'form': form,
        }
        return render(request, 'student/student_signup.html', context)
    else:
        return render(request, 'Error.html')


def student_edit(request, pk):
    if request.user.is_superuser:
        student = Student.objects.get(id=pk)
        form = StudentSignupForm(instance=student)
        if request.method == "POST":
            form = StudentSignupForm(request.POST, request.FILES, instance=student)
            if form.is_valid():
                form.save()
                messages.success(request, 'Student details edited successfully')
                return redirect('student-list-url')
        context = {
            'form': form,
            'student': student
        }
        return render(request, 'student/student_edit.html', context)
    else:
        return render(request, 'Error.html')


def student_status(request, pk):
    if request.user.is_superuser:
        student = Student.objects.get(id=pk)
        student.is_active = not student.is_active
        student.save()
        messages.success(request, 'Student active status changed to Inactive')
        return redirect('student-list-url')
    else:
        return render(request, 'Error.html')


def student_list(request):
    if request.user.is_superuser:
        list_choice = 0
        students = Student.objects.all()
        if request.method == 'POST':
            choice = request.POST.get('choice')
            if int(choice) == 0:
                return redirect('student-list-url')
            elif int(choice) == 1:
                return redirect('student-list-active-url')
            elif int(choice) == 2:
                return redirect('student-list-inactive-url')
        return render(request, 'student/student_list.html',
                      {
                          'students': students,
                          'list_choice': list_choice})
    else:
        return render(request, 'Error.html')


def student_list_active_only(request):
    if request.user.is_superuser:
        list_choice = 1
        students = Student.objects.filter(
            is_active=True
        )
        return render(request, 'student/student_list.html',
                      {
                          'students': students,
                          'list_choice': list_choice})
    else:
        return render(request, 'Error.html')


def student_list_inactive_only(request):
    if request.user.is_superuser:
        list_choice = 1
        students = Student.objects.filter(
            is_active=True
        )
        return render(request, 'student/student_list.html',
                      {
                          'students': students,
                          'list_choice': list_choice})
    else:
        return render(request, 'Error.html')


# ==============Student Views End====================================

# ==================Teaching clock in/out=========================================================
def calculate_duration_to_now(trainer):
    if trainer.clock_in and not trainer.clock_out_status:
        time_now = datetime.datetime.now()
        duration = time_now.astimezone() - trainer.clock_in
        minutes = int(duration.total_seconds() / 60)
        hours = minutes // 60
        rem_minutes = minutes % 60
        return f"{hours}hrs {rem_minutes}", hours
    return None


def calculate_teaching_hours(trainer, unit):
    now = t2.now()
    # start_of_week = now - timedelta(days=now.weekday())
    # start_of_week = make_aware(start_of_week)
    start_of_week = now - timedelta(days=now.weekday())  # Calculate the start of the current week (Monday)
    start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)  # Set the time to the start of the day
    start_of_week = make_aware(start_of_week)  
    trainer_unit = TrainerUnit.objects.get(trainer=trainer, unit=unit)
    teaching_attendances = TeachingAttendance.objects.filter(
        trainer_unit=trainer_unit,
        clock_in__gte=start_of_week,
        clock_in__lte=now
    )
    expected_hour = trainer_unit.unit.teaching_hrs_week
    if expected_hour:
        expected_hours = expected_hour
    else:
        expected_hours = 1
    total_minutes = sum(attendance.time_taken for attendance in teaching_attendances if attendance.time_taken)
    total_hours = int(total_minutes) // 60
    rem_minutes = total_minutes % 60
    percentage_hours_covered = (total_hours * 100) / int(expected_hours)
    remaining_hours = expected_hours - total_hours
    return f"{total_hours}Hrs {rem_minutes} Minutes", percentage_hours_covered, expected_hour, remaining_hours


def clock(request):
    if request.user.is_superuser:
        if request.method == 'POST':
            pf_number = request.POST.get('pf-number').upper()
            try:
                trainer = Trainer.objects.get(pf_number=pf_number)
                if trainer and trainer.is_active:
                    try:
                        t = TeachingAttendance.objects.get(
                            trainer_unit__trainer=trainer,
                            clock_in_status=True,
                            clock_out_status=False,
                        )
                        duration, hours = calculate_duration_to_now(t)
                        hours_covered_covered, percentage_hours_week, expected_hours_week, remaining_hours = calculate_teaching_hours(
                            trainer, t.trainer_unit.unit)
                        messages.error(request, 'Clock out of current Class ( ' + t.trainer_unit.unit.unit_name + ' )')
                        context = {
                            'hours_covered_covered': hours_covered_covered,
                            'percentage_hours_week': percentage_hours_week,
                            'expected_hours_week': expected_hours_week,
                            'remaining_hours': remaining_hours,
                            'details': t,
                            'duration': duration,
                            'hours': hours,
                        }
                        return render(request, 'clock/admin_clock_out.html', context)

                    except TeachingAttendance.DoesNotExist:
                        return HttpResponseRedirect('/clock/teaching/clockin/' + str(trainer.id))
                elif trainer and not trainer.is_active:
                    messages.error(request, trainer.name + ' Not allowed to clock in for a class')
            except Trainer.DoesNotExist:
                messages.error(request, 'Invalid PF number. Trainer not found')
        return render(request, 'clock/admin_clock_home.html')
    else:
        return render(request, 'Error.html')


def clock_in(request, pk):
    if request.user.is_superuser:
        units = Unit.objects.filter(trainerunit__trainer_id=pk)
        trainer = Trainer.objects.get(
            id=pk
        )

        if request.method == 'POST':
            unit_id = request.POST.get('unit_id')
            unit = Unit.objects.get(id=unit_id)
            time = datetime.datetime.now()
            t = TeachingAttendance.objects.create(
                trainer_unit=TrainerUnit.objects.get(
                    trainer=trainer,
                    unit=unit
                ),
                clock_in=time.astimezone(),
                clock_in_status=True,
            )
            t.save()
            messages.success(request,
                             'Clocked in for ( ' + unit.unit_name + ') Class at ' + str(time.time()))
            return redirect(clock)
        context = {
            'units': units,
            'trainer': trainer
        }
        return render(request, 'clock/admin_clock_in.html', context)
    else:
        return render(request, 'Error.html')


def calculate_duration(t_a):
    if t_a.clock_in and t_a.clock_out and t_a.clock_out_status:
        duration = t_a.clock_out - t_a.clock_in
        time_taken = duration.total_seconds() / 60
        return time_taken
    return None


def clock_out(request, pk):
    if request.user.is_superuser:
        if request.method == 'POST':
            roll = request.POST.get('roll')
            time = datetime.datetime.now()
            t = TeachingAttendance.objects.get(id=pk)
            t.roll = roll
            t.clock_out_status = True
            t.clock_out = time.astimezone()
            t.save()
            t_a = TeachingAttendance.objects.get(id=pk)
            time_taken = calculate_duration(t_a)
            if time_taken > t.trainer_unit.unit.teaching_hrs_per_class:
                t_a.time_taken = t.trainer_unit.unit.teaching_hrs_per_class
            else:
                t_a.time_taken = time_taken
            t_a.save()
            messages.success(request,
                             'Clocked out of ' + t_a.trainer_unit.unit.unit_name + ' class at ' + str(time.time()))
            return redirect('clock-url')
        context = {
        }
        return render(request, 'clock/admin_clock_out.html', context)
    else:
        return render(request, 'Error.html')


def clock_history(request):
    if request.user.is_superuser:
        if request.method == 'POST':
            pf_number = request.POST.get('pf-number').upper()
            date1 = request.POST.get('date1')
            date2 = request.POST.get('date2')
            start_date = t2.strptime(date1, '%Y-%m-%d')
            end_date = t2.strptime(date2, '%Y-%m-%d')

            DatePicker.start_date = start_date
            DatePicker.end_date = end_date
            try:
                trainer = Trainer.objects.get(pf_number=pf_number)
                if trainer:
                    clocks = TeachingAttendance.objects.filter(
                        Q(clock_in__date__gte=start_date) &
                        Q(clock_in__date__lte=end_date) &
                        Q(trainer_unit__trainer=trainer)
                    )
                    if clocks:
                        messages.success(request, 'Success!')
                        context = {
                            'clocks': clocks,
                            'trainer': trainer,
                            'start': start_date,
                            'end': end_date
                        }
                        return render(request, 'clock/admin_clock_history.html', context)
                    else:
                        messages.error(request, 'No clock information')
            except Trainer.DoesNotExist:
                messages.error(request, 'Invalid PF number. Trainer not found')
        return render(request, 'clock/admin_clock_history.html')
    else:
        return render(request, 'Error.html')


class DatePicker:
    date1 = ''
    date2 = ''
    start_date = ''
    end_date = ''


def clock_list(request):
    if request.user.is_superuser:
        if request.method == 'POST':
            date1 = request.POST.get('date1')
            date2 = request.POST.get('date2')

            DatePicker.date1 = date1
            DatePicker.date2 = date2
            if date1 and date2:
                try:
                    start_date = t2.strptime(date1, '%Y-%m-%d')
                    end_date = t2.strptime(date2, '%Y-%m-%d')
                    clocks = TeachingAttendance.objects.filter(
                        Q(clock_in__date__gte=start_date) & Q(clock_in__date__lte=end_date)
                    )
                    context = {
                        "clocks": clocks,
                        'start': start_date,
                        'end': end_date
                    }
                    return render(request, 'clock/admin_clock_history_list.html', context)
                except TeachingAttendance.DoesNotExist:
                    messages.error(request, 'No data found in that range')
                    return redirect('clock-list-url')
            else:
                messages.error(request, 'Please provide both start and end dates.')
                return redirect('clock-list-url')
        return render(request, 'clock/admin_clock_history_list.html')
    else:
        return render(request, 'Error.html')


def active_classes(request):
    if request.user.is_superuser:
        clocks = TeachingAttendance.objects.filter(
            Q(clock_out_status=False)
        )
        if not clocks:
            messages.error(request, 'No active classes')
            return redirect('clock-url')
        return render(request, 'clock/admin_clock_active_classes.html', {'clocks': clocks})
    else:
        return render(request, 'Error.html')


# Trainer clock in after face recognition=====================================
def face_recognition_view(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        image_data = data.get('image_data')
        image_bytes = base64.b64decode(image_data.split(',')[1])
        nparr = np.frombuffer(image_bytes, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        face_encodings = face_recognition.face_encodings(frame)

        if face_encodings:
            best_match = None
            lowest_distance = float('inf')
            for trainer in Trainer.objects.all():
                try:
                    known_encoding = np.fromstring(trainer.face_encoding, dtype=float, sep=',')
                    faceDistances = face_recognition.face_distance([known_encoding], face_encodings[0])
                    if faceDistances[0] < lowest_distance:
                        lowest_distance = faceDistances[0]
                        best_match = trainer
                except Exception as e:
                    print(e)
                    messages.error(request, 'Face not detected or trainer not registered')
            if best_match and lowest_distance <= 0.6:
                return JsonResponse({'success': True, 'trainer_name': best_match.name, 'id': best_match.id}, )
            else:
                return JsonResponse({'success': False, 'message': 'Face not recognized'})
        else:
            return JsonResponse({'success': False, 'message': 'No face detected'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request'})


def clock_face_recognition(request):
    return render(request, 'clock/clock_home_face.html')


def clock_after_face_recognition(request, pk):
    try:
        trainer = Trainer.objects.get(id=pk)
        if trainer:
            try:
                t = TeachingAttendance.objects.get(
                    trainer_unit__trainer=trainer,
                    clock_in_status=True,
                    clock_out_status=False,
                )
                duration, hours = calculate_duration_to_now(t)
                hours_covered_covered, percentage_hours_week, expected_hours_week, remaining_hours = calculate_teaching_hours(
                    trainer, t.trainer_unit.unit)
                context = {
                    'hours_covered_covered': hours_covered_covered,
                    'percentage_hours_week': percentage_hours_week,
                    'expected_hours_week': expected_hours_week,
                    'remaining_hours': remaining_hours,
                    'details': t,
                    'duration': duration,
                    'hours': hours,
                }
                return render(request, 'clock/clock_out.html', context)
            except TeachingAttendance.DoesNotExist:
                return HttpResponseRedirect('/clock/teaching/clockin/face/' + str(trainer.id))
    except Trainer.DoesNotExist:
        messages.error(request, 'Something. Try again')
    return render(request, 'clock/clock_home_face.html')


def clock_in_after_face_recognition(request, pk):
    units = Unit.objects.filter(trainerunit__trainer_id=pk)
    trainer = Trainer.objects.get(id=pk)
    if request.method == 'POST':
        pf_number = request.POST.get('pf_number').upper()
        hours_input = request.POST.get('hours')
        try:
            t = TeachingAttendance.objects.get(
                trainer_unit__trainer=trainer,
                clock_in_status=True,
                clock_out_status=False,
            )
            if t:
                duration, hours = calculate_duration_to_now(t)
                return render(request, 'clock/clock_out.html', {'details': t, 'duration': duration, 'hours': hours})
        except TeachingAttendance.DoesNotExist:
            try:
                if Trainer.objects.get(pf_number=pf_number):
                    unit_id = request.POST.get('unit_id')
                    unit = Unit.objects.get(id=unit_id)
                    time = datetime.datetime.now()
                    t = TeachingAttendance.objects.create(
                        trainer_unit=TrainerUnit.objects.get(trainer_id=pk, unit_id=unit_id),
                        clock_in=time.astimezone(),
                        clock_in_status=True,
                    )
                    t.save()
                    unit = Unit.objects.get(id=t.trainer_unit.unit.id)
                    teaching_hrs_per_class = unit.teaching_hrs_per_class
                    if hours_input:
                        hours_class = hours_input
                    elif not hours_input and teaching_hrs_per_class:
                        hours_class = teaching_hrs_per_class
                    unit.teaching_hrs_per_class = int(hours_class)
                    unit.save()
                    messages.success(request, 'Clocked in for ( ' + unit.unit_name + ') Class at ' + str(time.time()))
                    return redirect('clock-face-url')
            except Trainer.DoesNotExist:
                messages.error(request, 'Invalid or wrong PF/ID number')
                return redirect('clock-face-url')
    context = {
        'units': units,
        'trainer': trainer
    }
    return render(request, 'clock/clock_in.html', context)


def get_unit_data(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        unit_id = data.get('unit_id')
        trainer_id = data.get('trainer_id')
        if unit_id and trainer_id:
            trainer = Trainer.objects.get(id=trainer_id)
            print(trainer)
            try:
                unit = Unit.objects.get(id=unit_id)
                print(unit)
                hours_covered_covered, percentage_hours_week, expected_hours_week, remaining_hours = calculate_teaching_hours(
                    trainer, unit)
                context = {
                    'success': True,
                    'unit_name': unit.unit_name,
                    'unit_code': unit.unit_code,
                    'expected_hours': expected_hours_week,
                    'hours_covered': hours_covered_covered,
                    'percentage': percentage_hours_week,
                    'remaining_hours': remaining_hours,
                }
                print(context)
                return JsonResponse(context)
            except:
                return JsonResponse({'success': False, 'message': 'Unit Not Found'})
        else:
            return JsonResponse({'success': False, 'message': 'No Invalid Values Given'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request'})


def calculate_duration_after_face_recognition(t_a):
    if t_a.clock_in and t_a.clock_out and t_a.clock_out_status:
        duration = t_a.clock_out - t_a.clock_in
        time_taken = duration.total_seconds() / 60
        return time_taken
    return None


def clock_out_after_face_recognition(request, pk):
    if request.method == 'POST':
        pf_number = request.POST.get("pf_number").upper()
        try:
            if Trainer.objects.get(pf_number=pf_number):
                roll = request.POST.get('roll')
                time = datetime.datetime.now()
                t = TeachingAttendance.objects.get(id=pk)
                t.roll = roll
                t.clock_out = time.astimezone()
                t.clock_out_status = True
                t.save()
                t_a = TeachingAttendance.objects.get(id=pk)
                time_taken = calculate_duration_after_face_recognition(t_a)
                if time_taken > t.trainer_unit.unit.teaching_hrs_per_class:
                    t_a.time_taken = t.trainer_unit.unit.teaching_hrs_per_class
                else:
                    t_a.time_taken = time_taken
                t_a.save()
                messages.success(request,
                                 'Clocked out of ' + t.trainer_unit.unit.unit_name + ' class at ' + str(time.time()))
                return redirect('clock-face-url')
        except Trainer.DoesNotExist:
            messages.error(request, 'Invalid or wrong PF number')
            return redirect('clock-face-url')
    return render(request, 'clock/clock_out.html')


# Generate excel========================================================
@login_required(login_url='clock-face-url')
def generate_excel_for_clock_history(request, pk):
    if DatePicker.start_date != "" and DatePicker.end_date != "":
        trainer = Trainer.objects.get(id=pk)
        wb = Workbook()
        ws = wb.active
        ws.title = "Teaching Attendance"

        ws.append([f"TRAINER NAME: {trainer.name}"])
        ws.append([f"TRAINER PF NUMBER: {trainer.pf_number}"])
        ws.append(["UNIT NAME", "UNIT CODE", "CLASS DATE", "CLOCK IN TIME", "CLOCK OUT TIME", "ROLL", "DURATION"])
        clocks = TeachingAttendance.objects.filter(
            Q(clock_in__date__gte=DatePicker.start_date) &
            Q(clock_in__date__lte=DatePicker.end_date) &
            Q(trainer_unit__trainer=trainer) &
            Q(clock_out_status=True)
        )
        if clocks.exists():
            for clo in clocks:
                hours = clo.time_taken // 60
                minutes = clo.time_taken % 60
                time_taken = f"{hours} Hrs {minutes} Mins"
                row = [
                    clo.trainer_unit.unit.unit_name,
                    clo.trainer_unit.unit.unit_code,
                    clo.clock_in.date(),
                    str(clo.clock_in.astimezone().time()),
                    str(clo.clock_out.astimezone().time()),
                    clo.roll,
                    str(time_taken)
                ]
                ws.append(row)

            bold_font = Font(bold=True, condense=True)
            bold_font2 = Font(bold=True, condense=True)
            ws["A1"].font = bold_font
            ws["A2"].font = bold_font
            ws["A3"].font = bold_font2
            ws["B3"].font = bold_font2
            ws["C3"].font = bold_font2
            ws["D3"].font = bold_font2
            ws["E3"].font = bold_font2
            ws["F3"].font = bold_font2
            ws["G3"].font = bold_font2
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = (f'attachment; filename="{trainer.name}-{trainer.pf_number}-class'
                                               f'-attendance.xlsx"')
            with BytesIO() as stream:
                wb.save(stream)
                response.write(stream.getvalue())
            return response
        else:
            messages.error(request, 'No teaching attendance records found for this trainer.')
            return redirect('clock-history-url')
    else:
        return redirect('clock-history-url')


@csrf_exempt
@login_required(login_url='clock-face-url')
def generate_excel_for_clock_all(request):
    global minutes, hours
    date1 = DatePicker.date1
    date2 = DatePicker.date2
    if date1 and date2:
        try:
            start_date = t2.strptime(date1, '%Y-%m-%d')
            end_date = t2.strptime(date2, '%Y-%m-%d')
            wb = Workbook()
            ws = wb.active
            ws.title = "Teaching Attendance"
            ws.append([f"TEACHING ATTENDANCE BETWEEN {start_date} AND {end_date}"])
            ws.append(
                ["TRAINER NAME", "PF NUMBER", "UNIT NAME", "UNIT CODE", "CLASS DATE", "CLOCK IN TIME", "CLOCK OUT TIME",
                 "ROLL", "DURATION"])
            clocks = TeachingAttendance.objects.filter(
                Q(clock_in__date__gte=start_date) &
                Q(clock_in__date__lte=end_date) &
                Q(clock_out_status=True)
            )
            if clocks.exists():
                for clo in clocks:
                    if clo.time_taken:
                        hours = clo.time_taken // 60
                        minutes = clo.time_taken % 60
                    time_taken = f"{hours} Hrs {minutes} Mins"
                    row = [
                        clo.trainer_unit.trainer.name,
                        clo.trainer_unit.trainer.pf_number,
                        clo.trainer_unit.unit.unit_name,
                        clo.trainer_unit.unit.unit_code,
                        clo.clock_out.astimezone().date(),
                        clo.clock_in.astimezone().time(),
                        clo.clock_out.astimezone().time() if clo.clock_in_status else 'Not Clocked Out',
                        clo.roll,
                        str(time_taken)
                    ]
                    ws.append(row)

                bold_font = Font(bold=True, condense=True)
                bold_font2 = Font(bold=True, condense=True)
                ws["A1"].font = bold_font
                ws["A2"].font = bold_font2
                ws["B2"].font = bold_font2
                ws["C2"].font = bold_font2
                ws["D2"].font = bold_font2
                ws["E2"].font = bold_font2
                ws["F2"].font = bold_font2
                ws["G2"].font = bold_font2
                ws["H2"].font = bold_font2
                ws["I2"].font = bold_font2
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = (f'attachment; filename="class attendance between {start_date} '
                                                   f'and {end_date}.xlsx"')
                with BytesIO() as stream:
                    wb.save(stream)
                    response.write(stream.getvalue())
                return response
        except TeachingAttendance.DoesNotExist:
            messages.error(request, 'Invalid date format or No data found in that range')
            return redirect('clock-list-url')
    messages.error(request, 'Something went wrong')
    return redirect('clock-list-url')


@csrf_exempt
@login_required(login_url='clock-face-url')
def generate_excel_for_student_all(request):
    students = Student.objects.all()
    if students:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "EQUIP ALL STUDENTS (ACTIVE AND INACTIVE)"
            ws.append(["EQUIP ALL STUDENTS (ACTIVE AND INACTIVE)"])
            ws.append(["STUDENT NAME", "REG No.", "COURSE NAME", "PHONE", "EMAIL", "STATUS"])
            if students.exists():
                for student in students:
                    row = [
                        student.name,
                        student.registration_number,
                        student.course_taking,
                        student.phone,
                        student.email,
                        "ACTIVE" if student.is_active else "INACTIVE"
                    ]
                    ws.append(row)

                bold_font = Font(bold=True, condense=True)
                bold_font2 = Font(bold=True, condense=True)
                ws["A1"].font = bold_font
                ws["A2"].font = bold_font2
                ws["B2"].font = bold_font2
                ws["C2"].font = bold_font2
                ws["D2"].font = bold_font2
                ws["E2"].font = bold_font2
                ws["F2"].font = bold_font2
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="All Student In Equip list.xlsx"'
                with BytesIO() as stream:
                    wb.save(stream)
                    response.write(stream.getvalue())
                return response
        except TeachingAttendance.DoesNotExist:
            messages.error(request, 'No Student')
            return redirect('student-list-url')
    messages.error(request, 'Something went wrong')
    return redirect('student-list-url')


# Generate pdf================================================
def generate_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    styleH = styles['Heading1']
    styleN = styles['Normal']
    
    # Define custom paragraph style for justified text
    justified_style = ParagraphStyle(
        name='Justified',
        parent=styleN,
        alignment=4,  # Justify
    )
    
    # Title
    elements.append(Paragraph("Trainers and Their Units", styleH))
    
    # Table data
    data = [["NAME", "ID/PF NUMBER", "UNITS TEACHING"]]
    
    trainers = Trainer.objects.all()
    for trainer in trainers:
        units_teaching = TrainerUnit.objects.filter(trainer=trainer)
        units_list = [unit.unit.unit_name for unit in units_teaching]
        units_str = ", ".join(units_list)
        
        data.append([
            Paragraph(trainer.name, styleN),
            Paragraph(trainer.pf_number, styleN),
            Paragraph(units_str, justified_style)  # Use justified style here
        ])
    
    # Create table with column widths
    table = Table(data, colWidths=[100, 100, 300])
    
    # Style table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="trainers_units.pdf"'
    
    return response
# def generate_pdf(request):
#     buffer = BytesIO()
#     doc = SimpleDocTemplate(buffer, pagesize=A4)
#     elements = []
#     styles = getSampleStyleSheet()
#     styleH = styles['Heading1']
#     elements.append(Paragraph("Trainers and Their Units", styleH))
#     data = [["NAME", "ID/PF NUMBER", "UNITS TEACHING"]]
#     trainers = Trainer.objects.all()
#     for trainer in trainers:
#         units_teaching = TrainerUnit.objects.filter(trainer=trainer)
#         units_list = [unit.unit.unit_name for unit in units_teaching]
#         units_str = ", ".join(units_list)
#         data.append([trainer.name, trainer.pf_number, units_str])
#     table = Table(data)
#     table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
#         ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#         ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
#         ('GRID', (0, 0), (-1, -1), 1, colors.black),
#     ]))
    
#     elements.append(table)
#     doc.build(elements)
#     pdf = buffer.getvalue()
#     buffer.close()
#     response = HttpResponse(pdf, content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="trainers_units.pdf"'
#     return response


# def grouper(iterable, n):
#     args = [iter(iterable)] * n
#     return itertools.zip_longest(*args)


# def generate_pdf(request):
#     trainers = TrainerUnit.objects.all()

#     # Group TrainerUnits by Trainer
#     trainers_grouped = {}
#     for trainer_unit in trainers:
#         trainer = trainer_unit.trainer
#         if trainer not in trainers_grouped:
#             trainers_grouped[trainer] = []
#         trainers_grouped[trainer].append(trainer_unit.unit.unit_name)
#         trainers_grouped[trainer].append(trainer_unit.unit.unit_code)

#     # Prepare data for PDF generation
#     data = [("Name", "PF/ID No.", "Unit")]  # Sample header row
#     for trainer, units in trainers_grouped.items():
#         for unit in units:
#             data.append((trainer.name, trainer.pf_number, unit))

#     # Generate PDF
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment;' + 'filename=' + "Trainers-" + f"{str(t2.now())}.pdf"
#     c = canvas.Canvas(response, pagesize=portrait(A4))  # Landscape orientation
#     w, h = portrait(A4)
#     max_rows_per_page = 40
#     x_offset = 25
#     y_offset = 25
#     padding = 20
#     xlist = [x + x_offset for x in [0, 100, 200, 700, 800]]  # Adjust column positions as needed
#     ylist = [h - y_offset - i * padding for i in range(max_rows_per_page + 1)]

#     for rows in grouper(data, max_rows_per_page):
#         rows = tuple(filter(bool, rows))
#         c.grid(xlist, ylist[:len(rows) + 1])
#         for y, row in zip(ylist[:-1], rows):
#             for x, cell in zip(xlist, row):
#                 c.drawString(x + 2, y - padding + 3, str(cell))
#         c.showPage()
#     c.setFontSize(4.0, 4.0)
#     c.setTitle("TRAINERS AS AT " + f"{str(t2.now())}")
#     c.save()

#     return response
